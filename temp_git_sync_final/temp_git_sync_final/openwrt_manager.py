#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
OpenWrt 批量资产管理助手
A desktop tool for batch login detection of OpenWrt devices
"""

import sys
import asyncio
import csv
from typing import List, Dict, Tuple, Optional
from datetime import datetime

import aiohttp
from PyQt6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QSplitter, QVBoxLayout, QHBoxLayout,
    QPlainTextEdit, QSpinBox, QPushButton, QTableWidget, QTableWidgetItem,
    QLabel, QFileDialog, QMessageBox, QHeaderView, QMenu, QFrame, QSizePolicy
)
from PyQt6.QtCore import Qt, QTimer, pyqtSignal, QObject, QPropertyAnimation, QEasingCurve
from PyQt6.QtGui import QPalette, QColor, QPainter, QAction
import qasync

# 导入设计系统和 UI 组件
from design_system import theme, Colors, Typography, BorderRadius, Spacing
from ui_components import (
    ApplePrimaryButton, AppleSecondaryButton, AppleDestructiveButton,
    StatusBadge, AppleCard, SectionHeader, StatsCard, CircularProgress,
    AppleSmallButton
)
from macos_widgets import MacTitleBar, MacMenuBar, MacWindow


# ============================================================================
# 凭证配置 - 请在此处修改登录凭证组合
# ============================================================================
CREDENTIALS_LIST = [
    {"username": "root", "password": "password"},  # 默认优先
    {"username": "root", "password": "admin"},
    {"username": "admin", "password": "admin"},
    {"username": "ubnt", "password": "ubnt"},
    {"username": "root", "password": "123456"},
]

# ============================================================================
# 多版本兼容配置 - 支持不同的 OpenWrt 变体
# ============================================================================

# 常见的登录路径（按优先级排序）
LOGIN_PATHS = [
    "/cgi-bin/luci",           # 标准 LuCI (OpenWrt 官方)
    "/cgi-bin/luci/admin",     # 部分定制版本
    "/login",                  # 简化路径
]

# 表单字段名变体（按优先级排序）
FIELD_NAME_VARIANTS = [
    {"username": "luci_username", "password": "luci_password"},  # 标准 LuCI
    {"username": "username", "password": "password"},            # 简化版 (PandoraBox等)
    {"username": "auth_username", "password": "auth_password"},  # 认证版
]


# ============================================================================
# 呼吸灯组件
# ============================================================================
class BreathingLight(QWidget):
    """呼吸灯动画组件"""
    
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setFixedSize(20, 20)
        self._opacity = 0.3
        self._is_active = False
        self._color = QColor(128, 128, 128)  # 默认灰色
        
        # 动画
        self._animation = QPropertyAnimation(self, b"opacity")
        self._animation.setDuration(1500)  # 1.5秒一个周期
        self._animation.setStartValue(0.3)
        self._animation.setEndValue(1.0)
        self._animation.setEasingCurve(QEasingCurve.Type.InOutQuad)
        self._animation.setLoopCount(-1)  # 无限循环
        
    def paintEvent(self, event):
        """绘制圆形呼吸灯"""
        painter = QPainter(self)
        painter.setRenderHint(QPainter.RenderHint.Antialiasing)
        
        # 设置颜色和透明度
        color = QColor(self._color)
        color.setAlphaF(self._opacity)
        painter.setBrush(color)
        painter.setPen(Qt.PenStyle.NoPen)
        
        # 绘制圆形
        painter.drawEllipse(0, 0, self.width(), self.height())
        
    def set_active(self, active: bool):
        """设置激活状态"""
        self._is_active = active
        if active:
            self._color = QColor(0, 255, 0)  # 绿色
            self._animation.start()
        else:
            self._color = QColor(128, 128, 128)  # 灰色
            self._animation.stop()
            self._opacity = 0.3
        self.update()
        
    def get_opacity(self):
        return self._opacity
    
    def set_opacity(self, value):
        self._opacity = value
        self.update()
        
    opacity = property(get_opacity, set_opacity)


# ============================================================================
# 信号桥接器（用于线程安全的UI更新）
# ============================================================================
class SignalBridge(QObject):
    """信号桥接器，用于从异步函数安全地更新UI"""
    update_row = pyqtSignal(int, str, str, str, str)  # row, status, user, password, notes
    task_finished = pyqtSignal()


# ============================================================================
# 异步登录管理器
# ============================================================================
class LoginManager:
    """OpenWrt 登录探测管理器"""
    
    def __init__(self, max_concurrent: int = 50):
        self.max_concurrent = max_concurrent
        self.semaphore = asyncio.Semaphore(max_concurrent)
        self.signal_bridge = SignalBridge()
        self._tasks = []
        self._is_running = False
        
    async def detect_login(self, row: int, target_url: str) -> Dict:
        """
        探测单个 OpenWrt 设备的登录
        支持多种 OpenWrt 变体（标准 LuCI, PandoraBox, iStoreOS 等）
        
        Args:
            row: 表格行号
            target_url: 目标 URL
            
        Returns:
            检测结果字典
        """
        async with self.semaphore:
            # 如果停止，直接返回
            if not self._is_running:
                return {"status": "已停止", "user": "", "password": "", "notes": "用户中止"}
            
            # 规范化 URL
            if not target_url.startswith(("http://", "https://")):
                target_url = f"http://{target_url}"
            
            # 配置超时
            timeout = aiohttp.ClientTimeout(total=10)  # 增加到10秒以适应更多设备
            
            try:
                async with aiohttp.ClientSession(timeout=timeout) as session:
                    # 🔄 遍历不同的登录路径（支持多种OpenWrt变体）
                    for login_path in LOGIN_PATHS:
                        if not self._is_running:
                            break
                        
                        login_url = f"{target_url.rstrip('/')}{login_path}"
                        
                        # 🆕 步骤1: 先访问登录页面，初始化 Session 和获取 Cookie
                        try:
                            async with session.get(
                                login_url,
                                ssl=False,
                                allow_redirects=True
                            ) as init_response:
                                if init_response.status not in [200, 401, 403]:
                                    # 此路径不可用，尝试下一个
                                    continue
                                await init_response.read()
                        except Exception:
                            # 此路径失败，尝试下一个
                            continue
                        
                        # 🆕 步骤2: 尝试不同的凭证组合
                        for cred in CREDENTIALS_LIST:
                            if not self._is_running:
                                break
                                
                            username = cred["username"]
                            password = cred["password"]
                            
                            # 🆕 步骤3: 尝试不同的字段名变体
                            for field_variant in FIELD_NAME_VARIANTS:
                                if not self._is_running:
                                    break
                                
                                # 更新状态
                                self.signal_bridge.update_row.emit(
                                    row,
                                    f"尝试 {username}/***",
                                    "",
                                    "",
                                    f"路径:{login_path}"
                                )
                                
                                try:
                                    # 构建表单数据
                                    data = {
                                        field_variant["username"]: username,
                                        field_variant["password"]: password
                                    }
                                    
                                    # 发送 POST 登录请求
                                    async with session.post(
                                        login_url,
                                        data=data,
                                        ssl=False,
                                        allow_redirects=False
                                    ) as response:
                                        # 检查是否登录成功
                                        cookies = response.cookies
                                        has_sysauth = any('sysauth' in str(cookie.key) for cookie in cookies.values())
                                        is_redirect = response.status in [302, 303]
                                        
                                        if has_sysauth or is_redirect:
                                            # 🎉 登录成功！
                                            self.signal_bridge.update_row.emit(
                                                row,
                                                "登录成功",
                                                username,
                                                password,
                                                f"{login_path} | {field_variant['username']}"
                                            )
                                            return {
                                                "status": "登录成功",
                                                "user": username,
                                                "password": password,
                                                "notes": f"{login_path} | {field_variant['username']}"
                                            }
                                        
                                        # 如果不是404或500错误，说明路径正确但密码错误
                                        # 继续尝试下一个凭证，不再尝试其他字段名变体
                                        if response.status not in [404, 500]:
                                            break
                                
                                except asyncio.TimeoutError:
                                    continue
                                except aiohttp.ClientConnectorError:
                                    # 连接错误，不再尝试其他路径和凭证
                                    self.signal_bridge.update_row.emit(
                                        row,
                                        "无法连接",
                                        "",
                                        "",
                                        "网络不通或端口未开放"
                                    )
                                    return {
                                        "status": "无法连接",
                                        "user": "",
                                        "password": "",
                                        "notes": "网络不通或端口未开放"
                                    }
                                except Exception:
                                    # 其他错误，尝试下一个字段名变体
                                    continue
                    
                    # 所有凭证都失败
                    self.signal_bridge.update_row.emit(
                        row,
                        "登录失败",
                        "",
                        "",
                        "所有凭证均无效"
                    )
                    return {
                        "status": "登录失败",
                        "user": "",
                        "password": "",
                        "notes": "所有凭证均无效"
                    }
                    
            except asyncio.TimeoutError:
                self.signal_bridge.update_row.emit(
                    row,
                    "连接超时",
                    "",
                    "",
                    "请求超时（8秒）"
                )
                return {
                    "status": "连接超时",
                    "user": "",
                    "password": "",
                    "notes": "请求超时（8秒）"
                }
            except Exception as e:
                self.signal_bridge.update_row.emit(
                    row,
                    "检测异常",
                    "",
                    "",
                    str(e)[:50]
                )
                return {
                    "status": "检测异常",
                    "user": "",
                    "password": "",
                    "notes": str(e)[:50]
                }
                
    async def batch_detect(self, targets: List[str]):
        """批量探测"""
        self._is_running = True
        self._tasks = []
        
        # 创建所有任务
        for idx, target in enumerate(targets):
            task = asyncio.create_task(self.detect_login(idx, target))
            self._tasks.append(task)
        
        # 等待所有任务完成
        await asyncio.gather(*self._tasks, return_exceptions=True)
        
        # 发送完成信号
        self.signal_bridge.task_finished.emit()
        
    def stop(self):
        """停止所有任务"""
        self._is_running = False
        for task in self._tasks:
            if not task.done():
                task.cancel()


# ============================================================================
# 主窗口
# ============================================================================
class MainWindow(QMainWindow):
    """主窗口"""
    
    def __init__(self):
        super().__init__()
        self.login_manager = LoginManager(max_concurrent=50)
        self.init_ui()
        self.connect_signals()
        
    def init_ui(self):
        """初始化UI - v0.app完整复刻"""
        self.setWindowTitle("OpenWrt Manager v2.0")
        self.setGeometry(100, 100, 1200, 800)
        self.setWindowFlags(Qt.WindowType.FramelessWindowHint)  # 无边框窗口
        self.setAttribute(Qt.WidgetAttribute.WA_TranslucentBackground) # 透明背景
        
        # === 主容器（带圆角边框的外框）===
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        
        # 主布局 - 留出阴影空间
        main_container = QVBoxLayout(central_widget)
        main_container.setContentsMargins(20, 20, 20, 20)
        main_container.setSpacing(0)
        
        # === macOS窗口容器 ===
        self.mac_window = MacWindow()
        
        # 窗口内部布局
        window_layout = QVBoxLayout(self.mac_window)
        window_layout.setContentsMargins(0, 0, 0, 0)
        window_layout.setSpacing(0)
        
        # === 1. 标题栏（52px，带红绿灯）===
        titlebar = MacTitleBar("OpenWrt Manager", "v2.0")
        titlebar.closeClicked.connect(self.close)
        titlebar.minimizeClicked.connect(self.showMinimized)
        titlebar.maximizeClicked.connect(self.toggle_maximize)
        
        # 实现标题栏拖拽
        def mousePressEvent(event):
            if event.button() == Qt.MouseButton.LeftButton:
                self.dragPosition = event.globalPosition().toPoint() - self.frameGeometry().topLeft()
                event.accept()
        
        def mouseMoveEvent(event):
            if event.buttons() == Qt.MouseButton.LeftButton:
                self.move(event.globalPosition().toPoint() - self.dragPosition)
                event.accept()
                
        titlebar.mousePressEvent = mousePressEvent
        titlebar.mouseMoveEvent = mouseMoveEvent
        
        window_layout.addWidget(titlebar)
        
        # === 2. 菜单栏（28px）===
        menubar = MacMenuBar()
        window_layout.addWidget(menubar)
        
        # === 3. 主内容区（左右分栏）===
        content_splitter = QSplitter(Qt.Orientation.Horizontal)
        content_splitter.setHandleWidth(1)
        content_splitter.setStyleSheet(f"""
            QSplitter::handle {{
                background-color: {Colors.Light.BORDER_SECONDARY.name()};
            }}
        """)
        
        # ==== 3.1 左侧边栏（280px，灰色背景）====
        sidebar = QWidget()
        sidebar.setFixedWidth(Spacing.SIDEBAR_WIDTH)
        sidebar.setStyleSheet(f"""
            QWidget {{
                background-color: #f5f5f7;
                border-right: 1px solid #d1d1d6;
            }}
            /* 确保内部控件背景透明或正确 */
            QPlainTextEdit {{
                background-color: white;
                border: 1px solid #c7c7c7;
                border-radius: 6px;
                padding: 8px;
                color: #1d1d1f;
            }}
        """)
        
        sidebar_layout = QVBoxLayout(sidebar)
        sidebar_layout.setContentsMargins(16, 16, 16, 16)
        sidebar_layout.setSpacing(16)
        
        # 目标输入
        sidebar_layout.addWidget(SectionHeader("目标"))
        self.url_input = QPlainTextEdit()
        self.url_input.setPlaceholderText("输入 IP 地址，每行一个...")
        self.url_input.setMinimumHeight(150)
        sidebar_layout.addWidget(self.url_input)
        
        # 并发设置
        sidebar_layout.addWidget(SectionHeader("设置"))
        settings_row = QHBoxLayout()
        settings_row.setSpacing(10)
        
        # 并发数设置 (下拉菜单风格)
        self.concurrent_btn = AppleSecondaryButton("并发 50 ⌄")
        self.concurrent_btn.setFixedHeight(36)
        self.concurrent_btn.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed)
        
        # 并发菜单
        concurrent_menu = QMenu(self)
        concurrent_menu.setStyleSheet(f"""
            QMenu {{
                background-color: white;
                border: 1px solid {Colors.Light.BORDER_PRIMARY.name()};
                border-radius: 6px;
                padding: 4px;
            }}
            QMenu::item {{
                padding: 4px 12px;
                border-radius: 4px;
                color: {Colors.Light.TEXT_PRIMARY.name()};
            }}
            QMenu::item:selected {{
                background-color: {Colors.Light.BUTTON_PRIMARY.name()};
                color: white;
            }}
        """)
        
        for val in [50, 100, 200, 500]:
            action = QAction(f"{val} 线程", self)
            action.triggered.connect(lambda checked, v=val: self.update_concurrent(v))
            concurrent_menu.addAction(action)
            
        self.concurrent_btn.setMenu(concurrent_menu)
        settings_row.addWidget(self.concurrent_btn)
        
        # 开始按钮
        self.start_button = ApplePrimaryButton("开始探测")
        self.start_button.setFixedHeight(36)
        self.start_button.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed)
        settings_row.addWidget(self.start_button)
        
        sidebar_layout.addLayout(settings_row)
        
        # 移除多余的侧边栏导出按钮
        # 导出功能已移动到右上角工具栏
        
        # 概览统计 (v0.app 风格: 无卡片，仅顶部边框)
        # sidebar_layout.addSpacing(Spacing.LG) # 移除额外间距，由布局控制
        
        stats_container = QWidget()
        stats_container.setStyleSheet("""
            QWidget {
                background-color: transparent;
                border: none;
                border-top: 1px solid #d1d1d6;
            }
            QLabel {
                border: none; 
                background: transparent;
            }
        """)
        
        # 概览标题
        stats_header_layout = QVBoxLayout()
        stats_header_layout.setContentsMargins(0, 12, 0, 8) # 顶部padding对应 border-t
        stats_label = QLabel("概览")
        stats_label.setStyleSheet("color: #8e8e8e; font-size: 11px; font-weight: 600; text-transform: uppercase;")
        stats_header_layout.addWidget(stats_label)
        
        self.stats_layout = QVBoxLayout(stats_container)
        self.stats_layout.setContentsMargins(0, 0, 0, 0)
        self.stats_layout.setSpacing(6) # space-y-2 (8px approx)
        self.stats_layout.addLayout(stats_header_layout) # 将标题放入容器以便一起管理
        
        # 初始化统计标签 (Row layout: Justify Between)
        self.stat_labels = {}
        for key, color_code, label_text in [
            ("success", "#34c759", "成功"),
            ("failed", "#ff3b30", "失败"),
            ("waiting", "#ff9500", "等待中"),
        ]:
            row = QHBoxLayout()
            
            # 标签
            lbl = QLabel(label_text)
            lbl.setStyleSheet("color: #1d1d1f; font-size: 13px;")
            row.addWidget(lbl)
            
            row.addStretch()
            
            # 数值
            val = QLabel("0")
            val.setStyleSheet(f"color: {color_code}; font-weight: 600; font-size: 13px;")
            row.addWidget(val)
            
            self.stats_layout.addLayout(row)
            self.stat_labels[key] = val
        
        # 总计 (带顶部分隔线)
        total_separator = QFrame()
        total_separator.setFixedHeight(1)
        total_separator.setStyleSheet("background-color: #d1d1d6; margin-top: 8px;")
        self.stats_layout.addWidget(total_separator)
        
        row_total = QHBoxLayout()
        lbl_total = QLabel("总计")
        lbl_total.setStyleSheet("color: #1d1d1f; font-size: 13px; font-weight: 500;")
        row_total.addWidget(lbl_total)
        row_total.addStretch()
        val_total = QLabel("0")
        val_total.setStyleSheet("color: #1d1d1f; font-weight: 600; font-size: 13px;")
        row_total.addWidget(val_total)
        
        self.stats_layout.addLayout(row_total)
        self.stat_labels["total"] = val_total
            
        sidebar_layout.addWidget(stats_container)
        
        sidebar_layout.addStretch()
        
        # ==== 3.2 右侧主内容区（白色背景）====
        main_content = QWidget()
        main_content.setStyleSheet(f"background-color: {Colors.Light.BG_PRIMARY.name()};")
        
        content_layout = QVBoxLayout(main_content)
        content_layout.setContentsMargins(0, 0, 0, 0)
        content_layout.setSpacing(0)
        
        # 工具栏
        toolbar = QWidget()
        toolbar.setFixedHeight(40)
        toolbar.setStyleSheet(f"""
            background-color: {Colors.Light.BG_TABLEHEAD.name()};
            border-bottom: 1px solid {Colors.Light.BORDER_SECONDARY.name()};
        """)
        toolbar_layout = QHBoxLayout(toolbar)
        toolbar_layout.setContentsMargins(16, 0, 16, 0)
        
        toolbar_label = QLabel("结果")
        toolbar_label.setStyleSheet(f"""
            color: {Colors.Light.TEXT_SECONDARY.name()};
            font-size: 11px;
            font-weight: 600;
            text-transform: uppercase;
        """)
        toolbar_layout.addWidget(toolbar_label)
        toolbar_layout.addStretch()
        
        self.export_toolbar_btn = AppleSmallButton("导出")
        
        # 创建导出菜单
        export_menu = QMenu(self)
        export_menu.addAction("📄 CSV 格式", self.export_csv)
        export_menu.addAction("📊 Excel 格式", self.export_excel)
        export_menu.addAction("📋 JSON 格式", self.export_json)
        export_menu.addAction("📝 Markdown", self.export_markdown)
        export_menu.addSeparator()
        export_menu.addAction("🖨️ 打印预览", self.print_preview)
        
        self.export_toolbar_btn.setMenu(export_menu)
        toolbar_layout.addWidget(self.export_toolbar_btn)
        
        content_layout.addWidget(toolbar)
        
        content_layout.addWidget(toolbar)
        
        # === 4.0 空状态页面 (v0.app 风格) ===
        self.empty_state = QWidget()
        empty_layout = QVBoxLayout(self.empty_state)
        empty_layout.setAlignment(Qt.AlignmentFlag.AlignCenter)
        empty_layout.setSpacing(16)
        
        # 心跳图标 (修改为灰色，更精致，对应 Activity icon)
        icon_label = QLabel("⚡") # 暂时使用 unicode, 理想情况是 SVG
        icon_label.setStyleSheet(f"""
            font-size: 40px;
            color: #c7c7c7; 
            margin-bottom: 12px;
        """)
        icon_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        empty_layout.addWidget(icon_label)
        
        empty_title = QLabel("暂无结果")
        empty_title.setStyleSheet("color: #1d1d1f; font-size: 15px; font-weight: 500; margin-bottom: 4px;")
        empty_title.setAlignment(Qt.AlignmentFlag.AlignCenter)
        empty_layout.addWidget(empty_title)
        
        empty_desc = QLabel("输入目标地址并点击开始探测")
        empty_desc.setStyleSheet("color: #8e8e8e; font-size: 13px;")
        empty_desc.setAlignment(Qt.AlignmentFlag.AlignCenter)
        empty_layout.addWidget(empty_desc)
        
        title_label = QLabel("暂无结果")
        title_label.setStyleSheet(f"""
            font-size: 16px;
            font-weight: 600;
            color: {Colors.Light.TEXT_PRIMARY.name()};
        """)
        title_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        empty_layout.addWidget(title_label)
        
        desc_label = QLabel("输入目标地址并点击开始探测")
        desc_label.setStyleSheet(f"""
            font-size: 13px;
            color: {Colors.Light.TEXT_SECONDARY.name()};
        """)
        desc_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        empty_layout.addWidget(desc_label)
        
        content_layout.addWidget(self.empty_state)

        # 表格 (默认隐藏)
        self.result_table = QTableWidget()
        self.result_table.hide()
        
        self.result_table.setColumnCount(6)
        self.result_table.setHorizontalHeaderLabels([
            "#", "地址", "状态", "用户名", "密码", "详情"
        ])
        
        # 表格样式
        self.result_table.setShowGrid(False)
        self.result_table.setAlternatingRowColors(False) # 由design_system控制item及hover
        self.result_table.verticalHeader().setVisible(False)
        self.result_table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.result_table.setSortingEnabled(True)
        self.result_table.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self.result_table.setFrameShape(QFrame.Shape.NoFrame)
        
        # 列宽
        header = self.result_table.horizontalHeader()
        header.setSectionResizeMode(0, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)
        header.setSectionResizeMode(2, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(3, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(4, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(5, QHeaderView.ResizeMode.Stretch)
        header.setFixedHeight(28)
        
        content_layout.addWidget(self.result_table)
        
        # === 添加侧边栏和主内容到分割器 ===
        content_splitter.addWidget(sidebar)
        content_splitter.addWidget(main_content)
        content_splitter.setStretchFactor(0, 0)  # 固定宽度
        content_splitter.setStretchFactor(1, 1)  # 可伸缩
        
        window_layout.addWidget(content_splitter)
        main_container.addWidget(self.mac_window)
        
        # === 隐藏功能性控件（逻辑兼容）===
        self.concurrent_spinbox = QSpinBox()
        self.concurrent_spinbox.setMinimum(1)
        self.concurrent_spinbox.setMaximum(500)
        self.concurrent_spinbox.setValue(50)
        self.concurrent_spinbox.hide()
        
        # 呼吸灯
        self.breathing_light = QWidget() # 占位
        self.breathing_light.set_active = lambda x: None
        
    def update_concurrent(self, val):
        """更新并发数"""
        self.concurrent_spinbox.setValue(val)
        self.concurrent_btn.setText(f"并发 {val} ⌄")
        
    def connect_signals(self):
        """连接信号"""
        self.start_button.clicked.connect(self.on_start_stop_clicked)
        self.concurrent_spinbox.valueChanged.connect(self.on_concurrent_changed)
        
        # 连接异步信号
        self.login_manager.signal_bridge.update_row.connect(self.update_table_row)
        self.login_manager.signal_bridge.task_finished.connect(self.on_task_finished)
        
        # 连接表格右键菜单
        self.result_table.customContextMenuRequested.connect(self.show_table_context_menu)
        
    def on_concurrent_changed(self, value):
        """并发数改变"""
        self.login_manager.max_concurrent = value
        self.login_manager.semaphore = asyncio.Semaphore(value)
        
    def on_start_stop_clicked(self):
        """开始/停止按钮点击"""
        if self.start_button.text() == "▶ 开始探测":
            self.start_detection()
        else:
            self.stop_detection()
            
    def start_detection(self):
        """开始探测"""
        # 获取输入的 URL 列表
        urls = self.url_input.toPlainText().strip().split('\n')
        urls = [url.strip() for url in urls if url.strip()]
        
        if not urls:
            QMessageBox.warning(self, "警告", "请先输入至少一个目标地址！")
            return
        
        # 切换显示状态
        self.empty_state.hide()
        self.result_table.show()
        
        # 清空表格
        self.result_table.setRowCount(0)
        
        # 初始化表格
        self.result_table.setRowCount(len(urls))
        for idx, url in enumerate(urls):
            # 序号
            self.result_table.setItem(idx, 0, QTableWidgetItem(str(idx + 1)))
            # 目标地址
            self.result_table.setItem(idx, 1, QTableWidgetItem(url))
            # 状态
            self.result_table.setItem(idx, 2, QTableWidgetItem("待检测"))
            # 其他列
            self.result_table.setItem(idx, 3, QTableWidgetItem(""))
            self.result_table.setItem(idx, 4, QTableWidgetItem(""))
            self.result_table.setItem(idx, 5, QTableWidgetItem(""))
        
        # 更新UI状态
        self.start_button.setText("停止")
        # 按钮样式由 ApplePrimaryButton 自动管理
        self.breathing_light.set_active(True)
        self.url_input.setEnabled(False)
        self.concurrent_btn.setEnabled(False)
        self.export_toolbar_btn.setEnabled(False)
        
        # 更新统计
        if hasattr(self, 'stat_labels'):
             self.stat_labels['total'].setText(str(len(urls)))
             self.stat_labels['waiting'].setText(str(len(urls)))
             self.stat_labels['success'].setText("0")
             self.stat_labels['failed'].setText("0")
        
        # 启动异步任务
        asyncio.ensure_future(self.login_manager.batch_detect(urls))
        
    def stop_detection(self):
        """停止探测"""
        self.login_manager.stop()
        self.on_task_finished()
        
        # 更新统计
        success_count = self.count_status("登录成功")
        failed_count = self.count_status("登录失败")
        timeout_count = self.count_status("连接超时")
        
        self.update_statistics()
        
    def update_table_row(self, row: int, status: str, user: str, password: str, notes: str):
        """更新表格行（线程安全）"""
        # 更新状态
        self.result_table.setItem(row, 2, QTableWidgetItem(status))
        self.result_table.setItem(row, 3, QTableWidgetItem(user))
        self.result_table.setItem(row, 4, QTableWidgetItem(password))
        self.result_table.setItem(row, 5, QTableWidgetItem(notes))
        
        # 设置颜色
        if status == "登录成功":
            color = QColor(144, 238, 144)  # 浅绿色
        elif status in ["登录失败", "连接超时", "无法连接", "检测异常"]:
            color = QColor(255, 182, 193)  # 浅红色
        else:
            color = QColor(255, 255, 255)  # 白色
            
        for col in range(6):
            item = self.result_table.item(row, col)
            if item:
                item.setBackground(color)
                
        # 更新统计
        self.update_statistics()
        
    def update_statistics(self):
        """更新统计信息"""
        total = self.result_table.rowCount()
        success = self.count_status("登录成功")
        failed = self.count_status("登录失败")
        timeout = self.count_status("连接超时") + self.count_status("无法连接") + self.count_status("检测异常")
        pending = self.count_status("待检测") + self.count_status("等待中") + self.count_status("正在")
        
        if hasattr(self, 'stat_labels'):
            self.stat_labels['total'].setText(str(total))
            self.stat_labels['success'].setText(str(success))
            self.stat_labels['failed'].setText(str(failed))
            self.stat_labels['waiting'].setText(str(pending))
        
    def count_status(self, status: str) -> int:
        """统计指定状态的数量"""
        count = 0
        for row in range(self.result_table.rowCount()):
            item = self.result_table.item(row, 2)
            if item and status in item.text():
                count += 1
        return count
        
    def on_task_finished(self):
        """任务完成"""
        self.start_button.setText("开始探测")
        # 按钮样式由 ApplePrimaryButton 自动管理
        self.breathing_light.set_active(False)
        self.url_input.setEnabled(True)
        self.concurrent_btn.setEnabled(True)
        self.export_toolbar_btn.setEnabled(True)
        
        # 最终统计
        success_count = self.count_status("登录成功")
        failed_count = self.count_status("登录失败")
        timeout_count = self.count_status("连接超时")
        no_connect_count = self.count_status("无法连接")
        
        self.update_statistics()
        
        # 弹窗提示
        QMessageBox.information(
            self,
            "完成",
            f"探测完成！\n成功: {success_count} | 失败: {failed_count} | 超时: {timeout_count}"
        )
    
    def show_table_context_menu(self, pos):
        """显示表格右键菜单"""
        # 获取当前选中的行
        current_row = self.result_table.rowAt(pos.y())
        if current_row < 0:
            return
        
        # 创建菜单
        menu = QMenu(self)
        
        # 复制操作
        copy_row_action = QAction("📋 复制整行", self)
        copy_row_action.triggered.connect(lambda: self.copy_table_row(current_row))
        menu.addAction(copy_row_action)
        
        copy_address_action = QAction("📋 复制地址", self)
        copy_address_action.triggered.connect(lambda: self.copy_cell(current_row, 1))
        menu.addAction(copy_address_action)
        
        copy_credential_action = QAction("📋 复制凭证", self)
        copy_credential_action.triggered.connect(lambda: self.copy_credentials(current_row))
        menu.addAction(copy_credential_action)
        
        menu.addSeparator()
        
        # 其他操作
        retest_action = QAction("🔄 重新测试", self)
        retest_action.triggered.connect(lambda: self.retest_device(current_row))
        menu.addAction(retest_action)
        
        delete_action = QAction("🗑️ 删除", self)
        delete_action.triggered.connect(lambda: self.delete_row(current_row))
        menu.addAction(delete_action)
        
        # 显示菜单
        menu.exec(self.result_table.mapToGlobal(pos))
    
    def copy_table_row(self, row: int):
        """复制整行到剪贴板"""
        row_data = []
        for col in range(self.result_table.columnCount()):
            item = self.result_table.item(row, col)
            row_data.append(item.text() if item else "")
        
        # 使用制表符分隔
        QApplication.clipboard().setText("\t".join(row_data))
        self.stats_text.setText(f"✅ 已复制第 {row + 1} 行")
    
    def copy_cell(self, row: int, col: int):
        """复制单个单元格"""
        item = self.result_table.item(row, col)
        if item:
            QApplication.clipboard().setText(item.text())
            self.stats_text.setText(f"✅ 已复制: {item.text()}")
    
    def copy_credentials(self, row: int):
        """复制凭证（用户名/密码）"""
        username_item = self.result_table.item(row, 3)
        password_item = self.result_table.item(row, 4)
        
        if username_item and password_item:
            credentials = f"{username_item.text()}/{password_item.text()}"
            QApplication.clipboard().setText(credentials)
            self.stats_text.setText(f"✅ 已复制凭证: {credentials}")
    
    def retest_device(self, row: int):
        """重新测试单个设备"""
        address_item = self.result_table.item(row, 1)
        if address_item:
            # TODO: 实现单个设备重测功能
            self.stats_text.setText(f"🔄 准备重新测试: {address_item.text()}")
    
    def delete_row(self, row: int):
        """删除行"""
        reply = QMessageBox.question(
            self,
            "确认删除",
            f"确定要删除第 {row + 1} 行吗？",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )
        
        if reply == QMessageBox.StandardButton.Yes:
            self.result_table.removeRow(row)
            self.update_statistics()
            self.stats_text.setText(f"🗑️ 已删除第 {row + 1} 行")
    
    def export_csv(self):
        """导出为 CSV"""
        filename, _ = QFileDialog.getSaveFileName(
            self,
            "导出为 CSV",
            f"openwrt_results_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
            "CSV Files (*.csv)"
        )
        
        if not filename:
            return
        
        try:
            with open(filename, 'w', newline='', encoding='utf-8-sig') as csvfile:
                writer = csv.writer(csvfile)
                
                # 写入表头
                headers = []
                for col in range(self.result_table.columnCount()):
                    headers.append(self.result_table.horizontalHeaderItem(col).text())
                writer.writerow(headers)
                
                # 写入数据
                for row in range(self.result_table.rowCount()):
                    row_data = []
                    for col in range(self.result_table.columnCount()):
                        item = self.result_table.item(row, col)
                        row_data.append(item.text() if item else "")
                    writer.writerow(row_data)
            
            QMessageBox.information(self, "成功", f"已导出 {self.result_table.rowCount()} 行数据到:\n{filename}")
        except Exception as e:
            QMessageBox.critical(self, "错误", f"导出失败:\n{str(e)}")
    
    def export_excel(self):
        """导出为 Excel 格式"""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            QMessageBox.warning(
                self, 
                "缺少依赖", 
                "需要安装 openpyxl 库:\npip install openpyxl"
            )
            return
        
        filename, _ = QFileDialog.getSaveFileName(
            self,
            "导出为 Excel",
            f"openwrt_results_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            "Excel Files (*.xlsx)"
        )
        
        if not filename:
            return
        
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "探测结果"
            
            # 写入表头
            headers = []
            for col in range(self.result_table.columnCount()):
                headers.append(self.result_table.horizontalHeaderItem(col).text())
            ws.append(headers)
            
            # 设置表头样式
            for cell in ws[1]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            # 写入数据
            for row in range(self.result_table.rowCount()):
                row_data = []
                for col in range(self.result_table.columnCount()):
                    item = self.result_table.item(row, col)
                    row_data.append(item.text() if item else "")
                ws.append(row_data)
                
                # 根据状态设置颜色
                status_cell = ws.cell(row + 2, 3)  # 状态列
                if "成功" in status_cell.value:
                    for cell in ws[row + 2]:
                        cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                elif "失败" in status_cell.value or "超时" in status_cell.value:
                    for cell in ws[row + 2]:
                        cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            
            # 自动调整列宽
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
            
            wb.save(filename)
            QMessageBox.information(self, "成功", f"已导出到:\n{filename}")
        except Exception as e:
            QMessageBox.critical(self, "错误", f"导出失败:\n{str(e)}")
    
    def export_json(self):
        """导出为 JSON 格式"""
        import json
        
        filename, _ = QFileDialog.getSaveFileName(
            self,
            "导出为 JSON",
            f"openwrt_results_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json",
            "JSON Files (*.json)"
        )
        
        if not filename:
            return
        
        try:
            # 获取表头
            headers = []
            for col in range(self.result_table.columnCount()):
                headers.append(self.result_table.horizontalHeaderItem(col).text())
            
            # 构建数据
            data = []
            for row in range(self.result_table.rowCount()):
                row_dict = {}
                for col in range(self.result_table.columnCount()):
                    item = self.result_table.item(row, col)
                    row_dict[headers[col]] = item.text() if item else ""
                data.append(row_dict)
            
            # 写入文件
            with open(filename, 'w', encoding='utf-8') as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
            
            QMessageBox.information(self, "成功", f"已导出 {len(data)} 条记录到:\n{filename}")
        except Exception as e:
            QMessageBox.critical(self, "错误", f"导出失败:\n{str(e)}")
    
    def export_markdown(self):
        """导出为 Markdown 格式"""
        filename, _ = QFileDialog.getSaveFileName(
            self,
            "导出为 Markdown",
            f"openwrt_results_{datetime.now().strftime('%Y%m%d_%H%M%S')}.md",
            "Markdown Files (*.md)"
        )
        
        if not filename:
            return
        
        try:
            with open(filename, 'w', encoding='utf-8') as f:
                # 写入标题
                f.write("# OpenWrt 探测结果\n\n")
                f.write(f"**生成时间**: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n")
                
                # 写入统计
                total = self.result_table.rowCount()
                success = self.count_status("登录成功")
                failed = self.count_status("登录失败")
                
                f.write("## 统计概览\n\n")
                f.write(f"- **总设备数**: {total}\n")
                f.write(f"- **成功**: {success}\n")
                f.write(f"- **失败**: {failed}\n")
                f.write(f"- **成功率**: {(success/total*100):.1f}%\n\n")
                
                # 写入表格
                f.write("## 详细结果\n\n")
                
                # 表头
                f.write("| ")
                for col in range(self.result_table.columnCount()):
                    f.write(self.result_table.horizontalHeaderItem(col).text() + " | ")
                f.write("\n")
                
                # 分隔线
                f.write("| " + " | ".join(["---"] * self.result_table.columnCount()) + " |\n")
                
                # 数据行
                for row in range(self.result_table.rowCount()):
                    f.write("| ")
                    for col in range(self.result_table.columnCount()):
                        item = self.result_table.item(row, col)
                        text = item.text() if item else ""
                        f.write(text + " | ")
                    f.write("\n")
            
            QMessageBox.information(self, "成功", f"已导出到:\n{filename}")
        except Exception as e:
            QMessageBox.critical(self, "错误", f"导出失败:\n{str(e)}")
    
    def print_preview(self):
        """打印预览"""
        QMessageBox.information(
            self,
            "打印预览",
            "打印功能开发中...\n\n当前可以:\n1. 导出为 PDF: 使用浏览器打开 Markdown 导出\n2. 直接打印: 导出 Excel 后打印"
        )
    
    def export_to_csv(self):
        """导出为 CSV"""
        filename, _ = QFileDialog.getSaveFileName(
            self,
            "导出为 CSV",
            f"openwrt_results_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
            "CSV Files (*.csv)"
        )
        
        if not filename:
            return
        
        try:
            with open(filename, 'w', newline='', encoding='utf-8-sig') as csvfile:
                writer = csv.writer(csvfile)
                
                # 写入表头
                headers = []
                for col in range(self.result_table.columnCount()):
                    headers.append(self.result_table.horizontalHeaderItem(col).text())
                writer.writerow(headers)
                
                # 写入数据
                for row in range(self.result_table.rowCount()):
                    row_data = []
                    for col in range(self.result_table.columnCount()):
                        item = self.result_table.item(row, col)
                        row_data.append(item.text() if item else "")
                    writer.writerow(row_data)
            
            QMessageBox.information(self, "成功", f"已成功导出到:\n{filename}")
        except Exception as e:
            QMessageBox.critical(self, "错误", f"导出失败:\n{str(e)}")

    def toggle_maximize(self):
        """切换最大化"""
        if self.isMaximized():
            self.showNormal()
            # 恢复圆角
            self.mac_window.setStyleSheet(f"""
                QFrame {{
                    background-color: {Colors.Light.BG_SECONDARY.name()};
                    border: 1px solid {Colors.Light.BORDER_PRIMARY.name()};
                    border-radius: 10px;
                }}
            """)
            self.centralWidget().layout().setContentsMargins(20, 20, 20, 20)
        else:
            self.showMaximized()
            # 最大化时去除圆角
            self.mac_window.setStyleSheet(f"""
                QFrame {{
                    background-color: {Colors.Light.BG_SECONDARY.name()};
                    border: none;
                    border-radius: 0px;
                }}
            """)
            self.centralWidget().layout().setContentsMargins(0, 0, 0, 0)


# ============================================================================
# 主程序入口
# ============================================================================
def main():
    """主函数"""
    app = QApplication(sys.argv)
    
    # 设置应用样式
    app.setStyle("Fusion")
    
    # 设置 qasync 事件循环
    loop = qasync.QEventLoop(app)
    asyncio.set_event_loop(loop)
    
    # 创建主窗口
    window = MainWindow()
    window.show()
    
    # 运行事件循环
    with loop:
        loop.run_forever()


if __name__ == "__main__":
    main()
